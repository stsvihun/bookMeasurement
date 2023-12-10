import os
from openpyxl import load_workbook
import pandas as pd

def append_data_to_excel(height, width, spine, excel_file_path):
    # Sample data to append
    new_data = {
        'Height (mm)': [height],
        'Width (mm)': [width],
        'Spine (mm)': [spine]
    }

    # Create a DataFrame from the new data
    df_new = pd.DataFrame(new_data)

    # Load the existing workbook
    book = load_workbook(excel_file_path)

    # Access the active sheet
    sheet = book.active

    def find_first_clear_row(sheet):
        for row_num in range(1, sheet.max_row + 1):
            if all(sheet.cell(row=row_num, column=col_num).value is None for col_num in range(1, 5)):
                return row_num
        return sheet.max_row + 1  # If no clear row is found, return the next row

    # Determine the last row with data in the sheet
    last_row = find_first_clear_row(sheet)

    # Write the new data to the sheet starting from the last row
    for index, row in df_new.iterrows():
        for col_num, value in enumerate(row, start=2):
            sheet.cell(row=last_row + index, column=col_num, value=value)

    # Save the workbook
    book.save(excel_file_path)

    # Open the Excel file based on the operating system
    #open_excel_file(excel_file_path)


def open_excel_file(file_path):
    system = os.name
    if system == 'posix':  # Linux or MacOS
        os.system('open "{}"'.format(file_path))
    elif system == 'nt':  # Windows
        os.system('start excel "{}"'.format(file_path))
    else:
        print("Unsupported operating system")